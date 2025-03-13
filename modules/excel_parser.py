import pandas as pd
import openpyxl
import re

class ExcelParser:
    """Classe para análise e extração de dados de planilhas Excel"""
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.sheets = {}
        self.formulas = {}
        self.structure = {}
        
    def extract_structure(self):
        """Extrai a estrutura completa da planilha Excel"""
        xls = pd.ExcelFile(self.excel_path)
        
        for sheet_name in xls.sheet_names:
            # Lê a planilha
            df = pd.read_excel(xls, sheet_name=sheet_name)
            self.sheets[sheet_name] = df
            
            # Extrai estrutura de colunas e tipos de dados
            column_structure = []
            for col in df.columns:
                data_type = str(df[col].dtype)
                column_structure.append({
                    "name": col,
                    "type": data_type,
                    "nullable": df[col].isnull().any()
                })
            
            self.structure[sheet_name] = {
                "columns": column_structure,
                "rows": len(df)
            }
            
        return self.structure
    
    def extract_formulas(self):
        """Extrai as fórmulas da planilha usando openpyxl"""
        wb = openpyxl.load_workbook(self.excel_path, data_only=False)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_formulas = {}
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.formula:
                        sheet_formulas[cell.coordinate] = cell.formula
            
            self.formulas[sheet_name] = sheet_formulas
        
        return self.formulas
    
    def translate_formulas(self):
        """Traduz fórmulas Excel para código Python que pode ser usado na aplicação web"""
        python_formulas = {}
        
        for sheet_name, sheet_formulas in self.formulas.items():
            python_sheet_formulas = {}
            
            for cell, formula in sheet_formulas.items():
                # Converte a fórmula Excel para Python (simplificado para exemplo)
                python_formula = self._excel_to_python_formula(formula)
                python_sheet_formulas[cell] = python_formula
            
            python_formulas[sheet_name] = python_sheet_formulas
        
        return python_formulas
    
    def _excel_to_python_formula(self, excel_formula):
        """Converte uma fórmula Excel em código Python executável"""
        # Esta é uma implementação simplificada para demonstração
        
        # Substitui referências de células por acesso a dicionário
        # Exemplo: A1 se torna data['A1']
        cell_pattern = r'([A-Z]+[0-9]+)'
        python_formula = re.sub(cell_pattern, r"data['\1']", excel_formula)
        
        # Substitui funções Excel comuns por equivalentes Python
        replacements = {
            'SUM': 'sum',
            'AVERAGE': 'np.mean',
            'MAX': 'max',
            'MIN': 'min',
            'IF': 'lambda x, y, z: y if x else z',
            # Adicionar mais funções conforme necessário
        }
        
        for excel_func, python_func in replacements.items():
            python_formula = python_formula.replace(excel_func, python_func)
        
        return python_formula