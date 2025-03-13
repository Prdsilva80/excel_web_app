"""
Módulo para processamento de dados das planilhas
"""
import pandas as pd
import numpy as np
import json
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os

def processar_dados(dados_planilha, estrutura_planilha):
    """
    Processa os dados de uma planilha e prepara para exibição ou cálculos
    
    Args:
        dados_planilha (dict): Dados da planilha no formato JSON
        estrutura_planilha (dict): Estrutura da planilha no formato JSON
        
    Returns:
        dict: Dados processados e formatados
    """
    resultado = {}
    
    # Se não há dados, retorna vazio
    if not dados_planilha:
        return resultado
    
    # Converte dados para o tipo correto com base na estrutura
    for sheet_name, sheet_struct in estrutura_planilha.items():
        for coluna in sheet_struct.get('columns', []):
            nome_coluna = coluna['name']
            tipo_coluna = coluna['type']
            
            # Pula se a coluna não existe nos dados
            if nome_coluna not in dados_planilha:
                continue
                
            valor = dados_planilha[nome_coluna]
            
            # Converte para o tipo correto
            if 'int' in tipo_coluna:
                try:
                    resultado[nome_coluna] = int(valor) if valor else 0
                except (ValueError, TypeError):
                    resultado[nome_coluna] = 0
            elif 'float' in tipo_coluna:
                try:
                    resultado[nome_coluna] = float(valor) if valor else 0.0
                except (ValueError, TypeError):
                    resultado[nome_coluna] = 0.0
            elif 'bool' in tipo_coluna:
                resultado[nome_coluna] = bool(valor) if valor else False
            elif 'date' in tipo_coluna:
                if valor:
                    try:
                        # Tenta converter para data
                        if isinstance(valor, str):
                            # Tenta diferentes formatos de data
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
                                try:
                                    data = datetime.strptime(valor, fmt).date()
                                    resultado[nome_coluna] = data.isoformat()
                                    break
                                except ValueError:
                                    continue
                        else:
                            resultado[nome_coluna] = valor
                    except Exception:
                        resultado[nome_coluna] = valor
                else:
                    resultado[nome_coluna] = None
            else:
                # Strings e outros tipos
                resultado[nome_coluna] = str(valor) if valor else ""
    
    return resultado

def agrupar_dados(dados_list, coluna_agrupamento, colunas_soma=None, colunas_media=None):
    """
    Agrupa dados por uma coluna específica e calcula somas e médias para outras colunas
    
    Args:
        dados_list (list): Lista de dicionários com os dados
        coluna_agrupamento (str): Nome da coluna para agrupar
        colunas_soma (list): Lista de colunas para calcular soma
        colunas_media (list): Lista de colunas para calcular média
        
    Returns:
        dict: Dados agrupados com cálculos
    """
    if not dados_list:
        return {}
        
    # Converte para DataFrame
    df = pd.DataFrame(dados_list)
    
    # Define colunas para operações se não fornecidas
    if colunas_soma is None:
        colunas_soma = []
        for col in df.columns:
            if df[col].dtype in ['int64', 'float64']:
                colunas_soma.append(col)
    
    if colunas_media is None:
        colunas_media = colunas_soma.copy()
    
    # Agrupa e calcula
    resultado = {}
    grupos = df.groupby(coluna_agrupamento)
    
    for nome_grupo, grupo in grupos:
        resultado[nome_grupo] = {
            'contagem': len(grupo),
            'soma': {},
            'media': {}
        }
        
        for col in colunas_soma:
            if col in grupo.columns:
                resultado[nome_grupo]['soma'][col] = grupo[col].sum()
                
        for col in colunas_media:
            if col in grupo.columns:
                resultado[nome_grupo]['media'][col] = grupo[col].mean()
    
    return resultado

def exportar_dados(dados_list, formato='csv', headers=None):
    """
    Exporta dados para diferentes formatos
    
    Args:
        dados_list (list): Lista de dicionários com os dados
        formato (str): Formato de exportação ('csv', 'excel', 'json')
        headers (list): Lista de cabeçalhos (colunas) a incluir
        
    Returns:
        bytes: Dados no formato especificado
    """
    if not dados_list:
        return None
        
    # Converte para DataFrame
    df = pd.DataFrame(dados_list)
    
    # Filtra colunas se headers fornecido
    if headers:
        colunas = [col for col in headers if col in df.columns]
        df = df[colunas]
    
    # Exporta para o formato solicitado
    if formato == 'csv':
        output = io.StringIO()
        df.to_csv(output, index=False, quoting=csv.QUOTE_NONNUMERIC)
        return output.getvalue().encode('utf-8')
        
    elif formato == 'excel':
        output = io.BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        return output.getvalue()
        
    elif formato == 'json':
        return json.dumps(df.to_dict('records'), default=str).encode('utf-8')
        
    else:
        return None

def gerar_relatorio_excel(dados_list, titulo, sheet_name='Relatório'):
    """
    Gera um arquivo Excel formatado com os dados
    
    Args:
        dados_list (list): Lista de dicionários com os dados
        titulo (str): Título do relatório
        sheet_name (str): Nome da planilha
        
    Returns:
        bytes: Arquivo Excel em formato de bytes
    """
    # Cria um workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Adiciona título
    ws['A1'] = titulo
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')
    
    # Se não há dados, retorna workbook vazio
    if not dados_list:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    # Adiciona cabeçalhos
    headers = list(dados_list[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        
        # Ajusta largura da coluna
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(15, len(header) + 2)
    
    # Adiciona dados
    for row_idx, item in enumerate(dados_list, 4):
        for col_idx, header in enumerate(headers, 1):
            valor = item.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Adiciona bordas
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=3, max_row=len(dados_list) + 3, 
                            min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Salva o workbook
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def calcular_estatisticas(dados_list, colunas_numericas=None):
    """
    Calcula estatísticas básicas para colunas numéricas
    
    Args:
        dados_list (list): Lista de dicionários com os dados
        colunas_numericas (list): Lista de colunas numéricas para calcular estatísticas
        
    Returns:
        dict: Estatísticas por coluna
    """
    if not dados_list:
        return {}
        
    # Converte para DataFrame
    df = pd.DataFrame(dados_list)
    
    # Identifica colunas numéricas se não fornecidas
    if colunas_numericas is None:
        colunas_numericas = []
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                colunas_numericas.append(col)
    
    # Calcula estatísticas
    resultado = {}
    for col in colunas_numericas:
        if col in df.columns:
            stats = {
                'min': df[col].min(),
                'max': df[col].max(),
                'mean': df[col].mean(),
                'median': df[col].median(),
                'std': df[col].std(),
                'count': df[col].count(),
                'sum': df[col].sum()
            }
            resultado[col] = stats
    
    return resultado