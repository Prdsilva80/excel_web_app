from flask import Flask, request, jsonify, render_template, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timezone
import os
import json
import pandas as pd
import openpyxl
import numpy as np
from config import get_config

def limitar(valor, limite):
    """Limita uma sequência a um número específico de itens"""
    return valor[:limite]

# Inicialização da aplicação
app = Flask(__name__)
app.config.from_object(get_config())

app.jinja_env.filters['limitar'] = limitar

@app.context_processor
def inject_now():
    """Injeta a data atual em todos os templates"""
    return {'now': datetime.now(timezone.utc)}

# Criar pastas necessárias
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Inicialização do banco de dados
db = SQLAlchemy(app)

# Inicialização do sistema de login
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor, faça login para acessar esta página.'
login_manager.login_message_category = 'info'

# Modelos de banco de dados
class User(UserMixin, db.Model):
    """Modelo de usuário"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128))
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def set_password(self, password):
        """Define a senha criptografada"""
        self.password_hash = generate_password_hash(password)
        
    def check_password(self, password):
        """Verifica a senha"""
        return check_password_hash(self.password_hash, password)

class Planilha(db.Model):
    """Modelo de planilha"""
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    descricao = db.Column(db.Text)
    estrutura = db.Column(db.Text)  # JSON com a estrutura da planilha
    formulas = db.Column(db.Text)   # JSON com as fórmulas
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class DadosPlanilha(db.Model):
    """Modelo para dados inseridos pelos usuários nas planilhas"""
    id = db.Column(db.Integer, primary_key=True)
    planilha_id = db.Column(db.Integer, db.ForeignKey('planilha.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    dados = db.Column(db.Text)  # JSON com os dados inseridos
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relacionamentos
    planilha = db.relationship('Planilha', backref='dados')
    user = db.relationship('User', backref='dados_planilhas')

# Classe para processamento de planilhas Excel
class ExcelParser:
    """Classe para análise e extração de dados de planilhas Excel"""
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.sheets = {}
        self.formulas = {}
        self.structure = {}
        
    def extract_structure(self):
        """Extrai a estrutura completa da planilha Excel"""
        xls = pd.ExcelFile(self.excel_path)
        
        for sheet_name in xls.sheet_names:
            # Lê a planilha
            df = pd.read_excel(xls, sheet_name=sheet_name)
            self.sheets[sheet_name] = df
            
            # Extrai estrutura de colunas e tipos de dados
            column_structure = []
            for col in df.columns:
                data_type = str(df[col].dtype)
                column_structure.append({
                    "name": col,
                    "type": data_type,
                    "nullable": df[col].isnull().any()
                })
            
            self.structure[sheet_name] = {
                "columns": column_structure,
                "rows": len(df)
            }
            
        return self.structure
    
    def extract_formulas(self):
        """Extrai as fórmulas da planilha usando openpyxl"""
        wb = openpyxl.load_workbook(self.excel_path, data_only=False)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_formulas = {}
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.formula:
                        sheet_formulas[cell.coordinate] = cell.formula
            
            self.formulas[sheet_name] = sheet_formulas
        
        return self.formulas

# Helpers para o login
@login_manager.user_loader
def load_user(user_id):
    """Carrega o usuário a partir do ID"""
    return db.session.get(User, int(user_id))

# Verificação de extensões de arquivo permitidas
def allowed_file(filename):
    """Verifica se a extensão do arquivo é permitida"""
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Rotas de autenticação
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Rota de login"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            login_user(user, remember=True)
            next_page = request.args.get('next', url_for('dashboard'))
            flash('Login realizado com sucesso!', 'success')
            return redirect(next_page)
        
        flash('Email ou senha inválidos.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Rota de logout"""
    logout_user()
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('login'))

# Rotas principais
@app.route('/')
def index():
    """Rota inicial redirecionando para o login ou dashboard"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principal do usuário"""
    planilhas = Planilha.query.all()
    
    # Para cada planilha, verificar se o usuário já inseriu dados
    usuario_dados = {}
    for planilha in planilhas:
        ultimo_dado = DadosPlanilha.query.filter_by(
            planilha_id=planilha.id,
            user_id=current_user.id
        ).order_by(DadosPlanilha.created_at.desc()).first()
        
        usuario_dados[planilha.id] = ultimo_dado is not None
    
    return render_template('dashboard.html', 
                        planilhas=planilhas, 
                        usuario_dados=usuario_dados,
                        is_admin=current_user.is_admin)

@app.route('/planilha/<int:planilha_id>')
@login_required
def ver_planilha(planilha_id):
    """Visualização de uma planilha específica"""
    planilha = Planilha.query.get_or_404(planilha_id)
    estrutura = json.loads(planilha.estrutura)
    
    # Busca dados anteriores do usuário (se existirem)
    dados_anteriores = DadosPlanilha.query.filter_by(
        planilha_id=planilha_id, 
        user_id=current_user.id
    ).order_by(DadosPlanilha.created_at.desc()).first()
    
    dados = {}
    if dados_anteriores:
        dados = json.loads(dados_anteriores.dados)
    
    # Formatar os campos para exibição na planilha web
    campos = []
    
    # Pegar apenas a primeira planilha para simplificar
    primeiro_sheet = next(iter(estrutura))
    for coluna in estrutura[primeiro_sheet]['columns']:
        nome = coluna['name']
        tipo = coluna['type']
        
        # Determinar tipo de campo HTML com base no tipo de dados pandas
        input_type = 'text'
        if 'int' in tipo or 'float' in tipo:
            input_type = 'number'
        elif 'date' in tipo:
            input_type = 'date'
        elif 'bool' in tipo:
            input_type = 'checkbox'
        
        # Valor atual (se existir)
        valor = dados.get(nome, '')
        
        campos.append({
            'nome': nome,
            'tipo': input_type,
            'valor': valor
        })
    
    return render_template('planilha.html', 
                        planilha=planilha, 
                        campos=campos)

@app.route('/planilha/<int:planilha_id>/salvar', methods=['POST'])
@login_required
def salvar_planilha(planilha_id):
    """Salva os dados inseridos pelo usuário na planilha"""
    planilha = Planilha.query.get_or_404(planilha_id)
    dados = request.form.to_dict()
    
    # Aplicar fórmulas (simplificado para o exemplo)
    if planilha.formulas:
        formulas = json.loads(planilha.formulas)
        # Aqui teríamos o código para aplicar as fórmulas aos dados
    
    # Salvar os dados
    novo_registro = DadosPlanilha(
        planilha_id=planilha_id,
        user_id=current_user.id,
        dados=json.dumps(dados)
    )
    
    db.session.add(novo_registro)
    db.session.commit()
    
    flash('Dados salvos com sucesso!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/planilhas/<int:planilha_id>/excluir', methods=['POST'])
@login_required
def excluir_planilha(planilha_id):
    """Exclui uma planilha e todos os dados relacionados"""
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    planilha = Planilha.query.get_or_404(planilha_id)
    
    # Excluir todos os dados relacionados à planilha
    DadosPlanilha.query.filter_by(planilha_id=planilha.id).delete()
    
    # Excluir a planilha
    db.session.delete(planilha)
    db.session.commit()
    
    flash(f'Planilha {planilha.nome} excluída com sucesso!', 'success')
    return redirect(url_for('gerenciar_planilhas'))

@app.route('/relatorios')
@login_required
def relatorios():
    """Página de relatórios do usuário"""
    # Buscar todas as planilhas
    planilhas = Planilha.query.all()
    
    # Para cada planilha, buscar histórico de dados do usuário
    historico = {}
    
    for planilha in planilhas:
        dados_usuario = DadosPlanilha.query.filter_by(
            planilha_id=planilha.id,
            user_id=current_user.id
        ).order_by(DadosPlanilha.created_at.desc()).all()
        
        if dados_usuario:
            historico[planilha.id] = {
                'planilha': planilha,
                'dados': [
                    {
                        'id': d.id,
                        'data': d.created_at,
                        'dados': json.loads(d.dados)
                    } for d in dados_usuario
                ]
            }
    
    return render_template('relatorios.html', historico=historico)

@app.route('/relatorios/<int:dados_id>')
@login_required
def ver_relatorio(dados_id):
    """Visualização detalhada de um relatório específico"""
    dados = DadosPlanilha.query.get_or_404(dados_id)
    
    # Verificar se o dado pertence ao usuário atual
    if dados.user_id != current_user.id and not current_user.is_admin:
        flash('Você não tem permissão para acessar este relatório.', 'danger')
        return redirect(url_for('relatorios'))
    
    planilha = dados.planilha
    dados_json = json.loads(dados.dados)
    
    return render_template('ver_relatorio.html', 
                        planilha=planilha, 
                        dados=dados,
                        dados_json=dados_json)

# Rotas de administração
@app.route('/admin')
@login_required
def admin():
    """Dashboard administrativo"""
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    usuarios = User.query.all()
    planilhas = Planilha.query.all()
    
    return render_template('admin/dashboard.html', 
                        usuarios=usuarios, 
                        planilhas=planilhas)

@app.route('/admin/usuarios')
@login_required
def gerenciar_usuarios():
    """Página de gerenciamento de usuários"""
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    usuarios = User.query.all()
    return render_template('admin/usuarios.html', usuarios=usuarios)

@app.route('/admin/usuarios/novo', methods=['GET', 'POST'])
@login_required
def novo_usuario():
    """Página para criar novo usuário"""
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        is_admin = 'is_admin' in request.form
        
        # Verificar se usuário já existe
        existing_user = User.query.filter(
            (User.username == username) | (User.email == email)
        ).first()
        
        if existing_user:
            flash('Usuário ou email já cadastrado.', 'danger')
            return render_template('admin/novo_usuario.html')
        
        # Criar novo usuário
        user = User(username=username, email=email, is_admin=is_admin)
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash(f'Usuário {username} criado com sucesso!', 'success')
        return redirect(url_for('gerenciar_usuarios'))
    
    return render_template('admin/novo_usuario.html')

@app.route('/admin/usuarios/<int:user_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_usuario(user_id):
    """Página para editar usuário existente"""
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        
        # Verificar se o nome de usuário ou email já existe (excluindo o usuário atual)
        existing_user = User.query.filter(
            (User.username == username) | (User.email == email),
            User.id != user_id
        ).first()
        
        if existing_user:
            flash('Nome de usuário ou email já está em uso.', 'danger')
            return render_template('admin/editar_usuario.html', user=user)
        
        # Atualizar dados do usuário
        user.username = username
        user.email = email
        
        # Alterar senha apenas se for fornecida
        if request.form.get('password'):
            user.set_password(request.form['password'])
        
        user.is_admin = 'is_admin' in request.form
        
        db.session.commit()
        
        flash(f'Usuário {username} atualizado com sucesso!', 'success')
        return redirect(url_for('gerenciar_usuarios'))
    
    return render_template('admin/editar_usuario.html', user=user)

@app.route('/admin/usuarios/<int:user_id>/excluir', methods=['POST'])
@login_required
def excluir_usuario(user_id):
    """Exclui um usuário e todos os seus dados"""
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    
    # Verificar se está tentando excluir o próprio usuário
    if user.id == current_user.id:
        flash('Você não pode excluir seu próprio usuário.', 'danger')
        return redirect(url_for('gerenciar_usuarios'))
    
    # Excluir todos os dados relacionados ao usuário
    DadosPlanilha.query.filter_by(user_id=user.id).delete()
    
    # Excluir o usuário
    db.session.delete(user)
    db.session.commit()
    
    flash(f'Usuário {user.username} excluído com sucesso!', 'success')
    return redirect(url_for('gerenciar_usuarios'))

@app.route('/admin/planilhas')
@login_required
def gerenciar_planilhas():
    """Página de gerenciamento de planilhas"""
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    planilhas = Planilha.query.all()
    return render_template('admin/planilhas.html', planilhas=planilhas)

@app.route('/admin/planilhas/importar', methods=['GET', 'POST'])
@login_required
def importar_planilha():
    """Página para importar nova planilha"""
    if not current_user.is_admin:
        flash('Acesso restrito a administradores.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        if 'arquivo' not in request.files:
            flash('Nenhum arquivo selecionado', 'danger')
            return render_template('admin/importar_planilha.html')
        
        arquivo = request.files['arquivo']
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado', 'danger')
            return render_template('admin/importar_planilha.html')
        
        if arquivo and allowed_file(arquivo.filename):
            # Salvar o arquivo temporariamente
            filename = secure_filename(arquivo.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            arquivo.save(filepath)
            
            # Processar a planilha
            try:
                nome = request.form['nome']
                descricao = request.form['descricao']
                
                parser = ExcelParser(filepath)
                estrutura = parser.extract_structure()
                formulas = parser.extract_formulas()
                
                # Criar registro no banco
                nova_planilha = Planilha(
                    nome=nome,
                    descricao=descricao,
                    estrutura=json.dumps(estrutura),
                    formulas=json.dumps(formulas)
                )
                
                db.session.add(nova_planilha)
                db.session.commit()
                
                flash(f'Planilha {nome} importada com sucesso!', 'success')
                
                # Remover arquivo temporário após processamento
                os.remove(filepath)
                
                return redirect(url_for('gerenciar_planilhas'))
                
            except Exception as e:
                flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
                return render_template('admin/importar_planilha.html')
        else:
            flash('Formato de arquivo inválido. Use apenas planilhas Excel (.xlsx, .xls)', 'danger')
            return render_template('admin/importar_planilha.html')
    
    return render_template('admin/importar_planilha.html')

# Inicialização do banco de dados e usuário admin
def criar_tabelas_e_admin():
    """Cria tabelas e usuário admin"""
    with app.app_context():  # Isso é importante!
        db.create_all()
        
        # Criar usuário administrador se não existir
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                email='admin@example.com',
                is_admin=True
            )
            admin.set_password('admin123')  # Alterar para senha segura em produção
            db.session.add(admin)
            db.session.commit()
            app.logger.info('Usuário administrador criado')

# Inicialização
if __name__ == '__main__':
    criar_tabelas_e_admin()  # Chame a função antes de iniciar o app
    app.run(debug=True)